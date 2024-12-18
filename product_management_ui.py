import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
import random
import openpyxl
import time
import pyperclip
from threading import Thread
from PIL import Image, ImageTk
import io
import os
import gc
import shutil
import threading

class ProductManagerUI:
    def __init__(self, products, save_products):
        self.file_path = "products.xlsx"
        self.products = products if products else []
        self.save_products = self.save_to_excel
        self.current_input_step = 0
        self.current_product = {}
        self.filtered_products = []
        self.pasteboard_monitoring = False
        self.is_running = True
        self.monitor_thread = None
        self.root = tk.Toplevel()
        self.setup_styles()  # 设置样式
        self.init_gui()
        self.load_products()
        self.filtered_products = self.products.copy()
        self.check_clipboard_updates()

    def setup_styles(self):
        """设置UI样式"""
        # 定义颜色
        self.colors = {
            'primary': '#2196F3',  # 主色调
            'secondary': '#64B5F6',  # 次要色调
            'success': '#4CAF50',  # 成功色
            'warning': '#FFC107',  # 警告色
            'error': '#F44336',  # 错误色
            'background': '#F5F5F5',  # 背景色
            'text': '#212121',  # 文本色
            'light_text': '#757575'  # 浅色文本
        }

        # 创建自定义样式
        style = ttk.Style()
        style.theme_use('clam')

        # 配置Treeview样式
        style.configure("Treeview",
            background=self.colors['background'],
            foreground=self.colors['text'],
            fieldbackground=self.colors['background'],
            rowheight=30,
            font=('Microsoft YaHei UI', 10))
        
        style.configure("Treeview.Heading",
            background=self.colors['primary'],
            foreground="white",
            relief="flat",
            font=('Microsoft YaHei UI', 10, 'bold'))
        
        style.map("Treeview.Heading",
            background=[('active', self.colors['secondary'])])

        # 配置按钮样式
        style.configure("Primary.TButton",
            background=self.colors['primary'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10))
        
        style.map("Primary.TButton",
            background=[('active', self.colors['secondary'])])

        # 配置危险按钮样式
        style.configure("Danger.TButton",
            background=self.colors['error'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10))
        
        style.map("Danger.TButton",
            background=[('active', '#E57373')])

        # 配置标签样式
        style.configure("Info.TLabel",
            background=self.colors['background'],
            foreground=self.colors['text'],
            font=('Microsoft YaHei UI', 10))

        # 配置输入框样式
        style.configure("Custom.TEntry",
            fieldbackground="white",
            padding=(5, 5))

    def init_gui(self):
        self.root.title("产品管理系统")
        self.root.protocol("WM_DELETE_WINDOW", self.return_to_main_menu)
        
        # 设置窗口样式
        self.root.configure(bg=self.colors['background'])
        
        window_width = 1080
        window_height = 700
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        position_top = int((screen_height - window_height) / 2)
        position_left = int((screen_width - window_width) / 2)
        self.root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

        # 创建主框架
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 搜索区域
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))

        self.search_label = ttk.Label(search_frame, 
                                    text="搜索产品:",
                                    style="Info.TLabel")
        self.search_label.pack(side=tk.LEFT, padx=5)
        
        self.search_entry = ttk.Entry(search_frame, 
                                    width=50,
                                    style="Custom.TEntry",
                                    font=('Microsoft YaHei UI', 10))
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_entry.bind('<Return>', self.handle_search)
        self.search_entry.bind("<FocusIn>", self.stop_pasteboard_monitoring)
        self.search_entry.bind("<FocusOut>", self.start_pasteboard_monitoring)

        # Treeview区域
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # 创建Treeview的滚动条
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Treeview
        columns = ("Name", "Specs", "Cost", "Link", "Code", "LocationCode", "StockQuantity", "AddedDate")
        self.tree = ttk.Treeview(tree_frame, 
                                columns=columns, 
                                show='headings',
                                style="Treeview",
                                yscrollcommand=tree_scroll.set)
        
        tree_scroll.config(command=self.tree.yview)
        
        # 设置列宽和对齐方式
        column_widths = {
            "Name": 200,
            "Specs": 150,
            "Cost": 100,
            "Link": 200,
            "Code": 100,
            "LocationCode": 100,
            "StockQuantity": 100,
            "AddedDate": 100
        }

        # 设置中文表头
        column_headers = {
            "Name": "产品名称",
            "Specs": "规格",
            "Cost": "成本",
            "Link": "链接",
            "Code": "商品代码",
            "LocationCode": "库位码",
            "StockQuantity": "库存数量",
            "AddedDate": "添加日期"
        }

        for col in columns:
            self.tree.heading(col, text=column_headers[col])
            self.tree.column(col, width=column_widths[col], anchor='center')

        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind('<Double-1>', self.on_item_double_click)

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)

        clear_button = ttk.Button(button_frame, 
                                text="清空数据库", 
                                style="Danger.TButton",
                                command=self.clear_database)
        clear_button.pack(side=tk.LEFT, padx=5)

        delete_button = ttk.Button(button_frame, 
                                 text="删除选中项", 
                                 style="Danger.TButton",
                                 command=self.delete_selected_entries)
        delete_button.pack(side=tk.LEFT, padx=5)

        fix_code_button = ttk.Button(button_frame,
                                   text="修复商品代码",
                                   style="Primary.TButton",
                                   command=self.fix_missing_codes)
        fix_code_button.pack(side=tk.LEFT, padx=5)

        back_button = ttk.Button(button_frame, 
                               text="返回主菜单", 
                               style="Primary.TButton",
                               command=self.return_to_main_menu)
        back_button.pack(side=tk.LEFT, padx=5)

        # 输入区域
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=10)

        self.entry_label = ttk.Label(input_frame, 
                                   text="请输入产品名称:",
                                   style="Info.TLabel")
        self.entry_label.pack(pady=5)

        self.entry_field = ttk.Entry(input_frame, 
                                   width=50,
                                   style="Custom.TEntry",
                                   font=('Microsoft YaHei UI', 10))
        self.entry_field.pack(pady=5)
        self.entry_field.bind('<Return>', self.handle_entry)
        self.entry_field.bind("<FocusIn>", self.stop_pasteboard_monitoring)
        self.entry_field.bind("<FocusOut>", self.start_pasteboard_monitoring)

        # 状态标签
        self.status_label = ttk.Label(main_frame, 
                                    text="",
                                    style="Info.TLabel",
                                    wraplength=window_width-40)
        self.status_label.pack(fill=tk.X, pady=5)

    def load_products(self):
        try:
            # 清除现有数据
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # 读取Excel文件
            wb = openpyxl.load_workbook('products.xlsx')
            ws = wb.active
            
            # 空现有产品列表
            self.products.clear()
            
            # 从第二行开始读取数据（跳过表头）
            for row in range(2, ws.max_row + 1):
                values = []
                for col in range(1, 9):  # 读取8列数据
                    cell_value = ws.cell(row=row, column=col).value
                    # 格式化特殊列
                    if col == 3 and cell_value is not None:  # Cost列
                        try:
                            # 确保是数字并格式化
                            cell_value = float(cell_value)
                            cell_value = f"¥{cell_value:.2f}"
                        except (ValueError, TypeError):
                            cell_value = f"¥{0.00}"
                    elif col == 5 and cell_value is not None:  # Code列
                        try:
                            # 确保是整数
                            cell_value = int(float(cell_value))
                        except (ValueError, TypeError):
                            cell_value = 0
                    elif col == 7 and cell_value is not None:  # StockQuantity列
                        try:
                            # 确保是整数
                            cell_value = int(float(cell_value))
                        except (ValueError, TypeError):
                            cell_value = 0
                    values.append(cell_value if cell_value is not None else "")
                
                # 创建产品字典
                product = {
                    "Name": values[0],
                    "Specs": values[1],
                    "Cost": values[2],
                    "Link": values[3],
                    "Code": values[4],  # 已经是整数
                    "LocationCode": values[5],
                    "StockQuantity": values[6],  # 已经是整数
                    "AddedDate": values[7]
                }
                self.products.append(product)
                
                # 添加到树形视图
                self.tree.insert("", "end", values=values)
            
            wb.close()
            
        except Exception as e:
            print(f"Error loading products: {e}")
            messagebox.showerror("错误", f"加载产品数据时出错: {str(e)}")

    def save_to_excel(self):
        temp_files = []
        wb = None
        temp_dir = "temp_images"
        try:
            # 确保临时目录存在
            os.makedirs(temp_dir, exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 修正表头
            headers = ["Name", "Specs", "Cost", "Link", "Code", "LocationCode", "StockQuantity", "AddedDate"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, product in enumerate(self.products, 2):
                # 确保按正确的顺序写入数据
                values = [
                    product.get("Name", ""),
                    product.get("Specs", ""),
                    product.get("Cost", ""),
                    product.get("Link", ""),
                    product.get("Code", 0),  # 确保Code有默认值0
                    product.get("LocationCode", ""),
                    product.get("StockQuantity", 0),
                    product.get("AddedDate", "")
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # 根据列类型设置单元格格式和值
                    if col == 3:  # Cost列
                        if isinstance(value, str) and value.startswith('¥'):
                            try:
                                cell.value = float(value.replace('¥', '').strip())
                            except ValueError:
                                cell.value = 0.0
                        cell.number_format = '#,##0.00'
                    elif col == 5:  # Code列
                        try:
                            # 确保Code是整数
                            if isinstance(value, int):
                                cell.value = value
                            elif isinstance(value, str) and value.strip():
                                cell.value = int(float(value))
                            else:
                                cell.value = 0
                        except (ValueError, TypeError) as e:
                            print(f"Code转换错误: {e}, 使用默认值0")
                            cell.value = 0
                        cell.number_format = '0'
                    elif col == 7:  # StockQuantity列
                        try:
                            cell.value = int(float(str(value))) if value else 0
                        except (ValueError, TypeError):
                            cell.value = 0
                        cell.number_format = '0'
                    elif col == 6:  # LocationCode列
                        cell.value = str(value)
                        cell.number_format = '@'  # 文本格式
                    else:
                        cell.value = str(value)
                
                # 设置行高以适应图片
                ws.row_dimensions[row].height = 150
                
                # 处理图片
                try:
                    product_code = product.get("Code", 0)  # 使用默认值0
                    if product_code:
                        img_dir = os.path.join("product_images", str(product_code))
                        if os.path.exists(img_dir):
                            for i, img_file in enumerate(sorted(os.listdir(img_dir))):
                                if i >= 6:  # 最多显示6张图片
                                    break
                                if img_file.endswith(('.jpg', '.png')):
                                    img_path = os.path.join(img_dir, img_file)
                                    try:
                                        temp_path = os.path.join(temp_dir, f"temp_img_{row}_{i}_{int(time.time())}.png")
                                        with Image.open(img_path) as img:
                                            img.thumbnail((100, 100))
                                            img.save(temp_path)
                                            temp_files.append(temp_path)
                                            xl_img = openpyxl.drawing.image.Image(temp_path)
                                            col_letter = openpyxl.utils.get_column_letter(9 + i)
                                            xl_img.anchor = f"{col_letter}{row}"
                                            ws.add_image(xl_img)
                                    except Exception as e:
                                        print(f"Error processing image {img_path}: {e}")
                                        continue
                except Exception as e:
                    print(f"Error adding images for product {product_code}: {e}")
            
            # 保存前检查临时文件
            missing_files = [f for f in temp_files if not os.path.exists(f)]
            if missing_files:
                raise FileNotFoundError(f"Missing temporary files: {missing_files}")
            
            # 保存文件
            wb.save(self.file_path)
            return True
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")
            return False
        finally:
            if wb:
                wb.close()
            
            # 清理临时文件
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except Exception as e:
                    print(f"Error deleting temp file {temp_file}: {e}")
            
            gc.collect()

    def clear_pasteboard(self):
        """清空剪贴板"""
        try:
            pyperclip.copy("")
        except Exception as e:
            print(f"Error clearing clipboard: {e}")

    def check_clipboard_updates(self):
        """定时检查剪贴板更新"""
        try:
            if not self.is_running:
                return
            
            if not hasattr(self, 'root') or not self.root.winfo_exists():
                print("Window no longer exists, stopping clipboard updates")
                return
                
            if self.pasteboard_monitoring:
                try:
                    current_clipboard = pyperclip.paste()
                    if hasattr(self, 'prev_clipboard') and current_clipboard != self.prev_clipboard:
                        self.prev_clipboard = current_clipboard
                        if self.current_input_step in [0, 3]:  # Product Name or Link
                            self.safe_update_entry(current_clipboard)
                    elif not hasattr(self, 'prev_clipboard'):
                        self.prev_clipboard = current_clipboard
                except Exception as e:
                    print(f"Error checking clipboard: {e}")
            
            # 如果程序仍在运行且窗口在，继续定时检查
            if self.is_running and hasattr(self, 'root') and self.root.winfo_exists():
                self.after_id = self.root.after(500, self.check_clipboard_updates)
        except Exception as e:
            print(f"Error in check_clipboard_updates: {e}")

    def monitor_pasteboard(self):
        """启动剪贴板监控"""
        self.prev_clipboard = pyperclip.paste()
        # 不再需要单独的监控线程，使用check_clipboard_updates代替

    def safe_update_entry(self, text):
        """在主线程中安全地更新输入框和处理入"""
        try:
            if not self.is_running:
                return
            self.entry_field.delete(0, tk.END)
            self.entry_field.insert(0, text)
            # 自动触发Enter键事件
            self.root.after(100, lambda: self.handle_entry(None))
        except Exception as e:
            print(f"Error in safe_update_entry: {e}")

    def stop_pasteboard_monitoring(self, event=None):
        """停止剪贴板监控"""
        print("Stopping pasteboard monitoring")
        self.pasteboard_monitoring = False

    def start_pasteboard_monitoring(self, event=None):
        """启动剪贴板监控"""
        print("Starting pasteboard monitoring")
        if not self.pasteboard_monitoring and self.is_running:
            self.pasteboard_monitoring = True
            self.clear_pasteboard()
            self.monitor_pasteboard()  # 现在只是初始化prev_clipboard

    def handle_search(self, event=None):
        search_query = self.search_entry.get().strip().lower()

        if search_query:
            self.filtered_products = [
                product for product in self.products if search_query in product["Name"].lower()
            ]
        else:
            self.filtered_products = self.products  # Show all products if search is empty

        self.refresh_tree()

    def refresh_tree(self):
        """刷新树形视图的显示"""
        try:
            self.tree.delete(*self.tree.get_children())
            for product in self.filtered_products:
                try:
                    # 获取商品代码并确保是整数
                    product_code = product.get("Code", 0)
                    if isinstance(product_code, str):
                        product_code = int(float(product_code))
                    elif isinstance(product_code, float):
                        product_code = int(product_code)
                    
                    values = [
                        product.get("Name", ""),
                        product.get("Specs", ""),
                        product.get("Cost", ""),
                        product.get("Link", ""),
                        str(product_code),  # 将商品代码转换为字符串显示
                        product.get("LocationCode", ""),
                        str(product.get("StockQuantity", 0)),  # 确保库存显示为字符串
                        product.get("AddedDate", "")
                    ]
                    self.tree.insert("", tk.END, values=tuple(values))
                except Exception as e:
                    print(f"处理商品时出错: {e}")
                    print(f"问题商品数据: {product}")
                    continue
        except Exception as e:
            print(f"刷新显示时出错: {e}")
            messagebox.showerror("错误", f"刷新显示时出错: {str(e)}")

    def on_column_double_click(self, event):
        # Identify the region where the double-click occurred
        region = self.tree.identify('region', event.x, event.y)
        
        if region == 'heading':  # Only trigger for column header
            # Identify the column by position
            col = self.tree.identify_column(event.x)  # This returns something like '#1', '#2', ...
            col = col.lstrip('#')  # Remove the '#' prefix from the column identifier
            self.adjust_column_width_on_double_click(col)  # Call function to adjust column width
    
    def adjust_column_width_on_double_click(self, col):
        # Get the column index
        col_index = int(col) - 1  # The identify_column method returns 1-based index (e.g., '#1' is index 1)
        
        # Get the maximum width based on the values in that column
        max_width = max(len(str(self.tree.item(child)['values'][col_index])) 
                        for child in self.tree.get_children())
        
        # Consider the header length as well
        max_width = max(max_width, len(self.tree.heading(self.tree['columns'][col_index])['text']))
        
        # Calculate new width (adjust multiplier if necessary)
        new_width = max_width * 10  # You can adjust the multiplier (10) based on your preference
        
        # Set the new column width
        self.tree.column(self.tree['columns'][col_index], width=new_width)
    def on_item_double_click(self, event):
        # Identify the region clicked
        region = self.tree.identify('region', event.x, event.y)
        
        if region == 'heading':
            # If the double-click happened on the column header, adjust the column width
            col = self.tree.identify_column(event.x)
            col = col.lstrip('#')  # Remove the '#' prefix
            self.adjust_column_width_on_double_click(col)
        elif region == 'cell':
            # If the double-click happened on a cell, edit the item
            item = self.tree.selection()  # Get the selected row/item
            if not item:
                return
            
            # Get the column index of the clicked cell
            col = self.tree.identify_column(event.x)  # Get the column identifier
            col_index = int(col.lstrip('#')) - 1  # Convert to 0-based index
            
            # Get the current value of the clicked cell
            item_values = self.tree.item(item[0], 'values')
            current_value = item_values[col_index]
            
            # Create an entry widget at the position of the clicked cell
            entry = ttk.Entry(self.root)  # 使用self.root作为父窗口
            entry.insert(0, current_value)
            
            # 获取树形视图中单元格的坐标
            bbox = self.tree.bbox(item[0], col)
            if bbox:
                # 将树形视图的坐标转换为root窗口的坐标
                x = self.tree.winfo_rootx() + bbox[0]
                y = self.tree.winfo_rooty() + bbox[1]
                w = bbox[2]
                
                # 设置Entry的位置和大小
                entry.place(in_=self.tree, x=bbox[0], y=bbox[1], width=w)
            
            # Function to save the edited value
            def save_edit(event=None):
                new_value = entry.get()
                if new_value != current_value:
                    # Update the product data in the backend (product list)
                    self.update_product(item[0], col_index, new_value)
                    
                entry.destroy()  # Remove the entry widget after editing
                self.refresh_tree()  # Refresh Treeview to show the updated value
            
            # Bind the entry widget to save the edit when pressing Enter
            entry.bind('<Return>', save_edit)
            
            # Bind the entry widget to save the edit when losing focus (clicking outside)
            entry.bind('<FocusOut>', save_edit)
            
            entry.focus_set()  # Focus on the entry widget
    
    def update_product(self, item_id, col_index, new_value):
        # Get the corresponding product data based on the item (row) in the Treeview
        item_values = self.tree.item(item_id, 'values')
        code = int(item_values[4])  # 将Code转换为整数
        
        # Find the product in the product list by matching the code
        for product in self.products:
            if int(product['Code']) == code:  # 确保比较时是整数
                column_name = self.tree['columns'][col_index]
                # 根据列类型处理新值
                if column_name == "Code":
                    try:
                        new_value = int(float(new_value))  # 确保Code是整数
                    except (ValueError, TypeError):
                        messagebox.showerror("错误", "Code必须是数字")
                        return
                elif column_name == "Cost":
                    try:
                        if isinstance(new_value, str) and new_value.startswith('¥'):
                            new_value = float(new_value.replace('¥', '').strip())
                        else:
                            new_value = float(new_value)
                        new_value = f"¥{new_value:.2f}"
                    except (ValueError, TypeError):
                        messagebox.showerror("错误", "价格必须是数字")
                        return
                elif column_name == "StockQuantity":
                    # 不允许在这里修改StockQuantity
                    messagebox.showinfo("提示", "库存数量只能在库存管理界面修改")
                    return
                elif column_name == "LocationCode":
                    # LocationCode 持为字符串，不需要数字验证
                    new_value = str(new_value).strip()
                
                # Update the correct field in the product dictionary
                product[column_name] = new_value
                break
        
        # Save the updated product list to Excel
        self.save_to_excel()

    def generate_unique_code(self):
        """生成唯一的7位整数商品代码"""
        max_attempts = 100  # 最大尝试次数
        for attempt in range(max_attempts):
            try:
                new_code = random.randint(1000000, 9999999)  # 生成7位整数
                # 检查现有产品列表
                existing_codes = set()
                for product in self.products:
                    try:
                        code = product.get("Code", 0)
                        if isinstance(code, str):
                            code = int(float(code))
                        existing_codes.add(code)
                    except (ValueError, TypeError):
                        continue
                
                # 如果生成的代码不在现有代码中
                if new_code not in existing_codes:
                    print(f"成功生成新Code: {new_code}")
                    return new_code
            except Exception as e:
                print(f"生成Code时出错 (尝试 {attempt + 1}/{max_attempts}): {e}")
                continue
        
        # 如果达到最大尝试次数仍未生成有效代码
        raise ValueError("无法生成唯一的商品代码，请重试")

    def handle_entry(self, event):
        try:
            # 确保在主线程中执行
            if threading.current_thread() is not threading.main_thread():
                self.root.after(0, lambda: self.handle_entry(event))
                return

            user_input = self.entry_field.get().strip()
            if self.current_input_step == 0:  # Product Name
                if user_input:
                    # 初始化新产品时，创建完整的数据结构
                    self.current_product = {
                        "Name": user_input,  # 直接设置名称
                        "Specs": "",
                        "Cost": "¥0.00",
                        "Link": "",
                        "Code": 0,  # 初始化为整数0
                        "LocationCode": "",
                        "StockQuantity": 0,
                        "AddedDate": datetime.now().strftime("%Y-%m-%d")
                    }
                    self.current_input_step += 1
                    self.entry_label.config(text="Enter Product Specifications (or press Enter to skip):")
                    self.entry_field.delete(0, tk.END)
                    self.stop_pasteboard_monitoring()
                else:
                    self.status_label.config(text="Error: Product name cannot be empty.")
            elif self.current_input_step == 1:  # Specifications
                self.current_product["Specs"] = user_input  # 可以为空
                self.current_input_step += 1
                self.entry_label.config(text="Enter Product Cost:")
                self.entry_field.delete(0, tk.END)
                self.stop_pasteboard_monitoring()
            elif self.current_input_step == 2:  # Cost
                try:
                    cost = float(user_input) if user_input else 0.0
                    self.current_product["Cost"] = f"¥{cost:.2f}"
                    self.current_input_step += 1
                    self.entry_label.config(text="Enter Product Link:")
                    self.entry_field.delete(0, tk.END)
                    self.start_pasteboard_monitoring()
                except ValueError:
                    self.status_label.config(text="Error: Please enter a valid cost.")
            elif self.current_input_step == 3:  # Link
                if user_input:
                    try:
                        print("\n开始处理新产品数据...")
                        
                        # 1. 保存链接
                        self.current_product["Link"] = user_input
                        print(f"已保存链接: {user_input}")
                        
                        # 2. 生成新的Code
                        try:
                            new_code = self.generate_unique_code()
                            if not isinstance(new_code, int) or new_code <= 0:
                                raise ValueError(f"Invalid code generated: {new_code}")
                            print(f"生成的新Code: {new_code} (类型: {type(new_code)})")
                            self.current_product["Code"] = new_code
                        except Exception as code_error:
                            print(f"生成Code时���错: {code_error}")
                            raise ValueError("无法生成有效的商品代码")
                        
                        # 3. 打印完整的产品数据进行验证
                        print("\n产品数据验证:")
                        for field in ["Name", "Specs", "Cost", "Link", "Code", "LocationCode", "StockQuantity", "AddedDate"]:
                            print(f"{field}: {type(self.current_product[field])} = {self.current_product[field]}")
                        
                        # 4. 保存到数据库
                        print("\n保存到数据库...")
                        new_product = self.current_product.copy()
                        self.products.append(new_product)
                        
                        # 5. 保存到Excel
                        print("保存到Excel...")
                        save_result = self.save_to_excel()
                        if not save_result:
                            raise Exception("保存到Excel失败")
                        
                        # 6. 更新显示
                        self.filtered_products = self.products.copy()
                        self.refresh_tree()
                        
                        # 7. 重置状态
                        print("重���状态...")
                        self.current_input_step = 0
                        self.current_product = {}
                        self.entry_label.config(text="Enter Product Name:")
                        self.entry_field.delete(0, tk.END)
                        self.start_pasteboard_monitoring()
                        
                        # 8. 更新状态标签
                        self.status_label.config(text="Product added successfully.")
                        print("产品添加完成!\n")
                        
                    except Exception as e:
                        print(f"\n错误详情:")
                        print(f"- 错误类型: {type(e)}")
                        print(f"- 错误信息: {str(e)}")
                        print(f"- 当前产品数据: {self.current_product}")
                        self.status_label.config(text=f"Error: {str(e)}")
                else:
                    self.status_label.config(text="Error: Product link cannot be empty.")
        except Exception as e:
            print(f"Error in handle_entry: {e}")
            self.status_label.config(text=f"Error: {str(e)}")

    def delete_selected_entries(self):
        selected_items = self.tree.selection()
        if selected_items:
            # 强制进行垃圾回收，释��可能的文件句柄
            gc.collect()
            
            product_codes_to_delete = []
            for item in selected_items:
                item_values = self.tree.item(item, 'values')
                if len(item_values) >= 5:
                    try:
                        # 获取商品代码（第5列）
                        product_code = item_values[4]
                        print(f"处理商品代码: {product_code}")
                        
                        # 如果商品代码不为空，则处理删除
                        if product_code and str(product_code).strip():
                            # 将商品代码转换为整数
                            product_code = int(float(str(product_code)))
                            product_codes_to_delete.append(product_code)
                            print(f"将删除商品代码: {product_code}")
                            
                            # 删除关联的图片文件
                            img_dir = os.path.join("product_images", str(product_code))
                            if os.path.exists(img_dir):
                                try:
                                    # 先尝试删除文件夹内的所有文件
                                    for img_file in os.listdir(img_dir):
                                        file_path = os.path.join(img_dir, img_file)
                                        try:
                                            if os.path.isfile(file_path):
                                                os.unlink(file_path)
                                            elif os.path.isdir(file_path):
                                                shutil.rmtree(file_path, ignore_errors=True)
                                        except Exception as e:
                                            print(f"删除文件出错 {file_path}: {e}")
                                            continue
                                    
                                    # 等待一段时间保文件被释放
                                    time.sleep(0.5)
                                    gc.collect()
                                    
                                    # 然后尝试删除文件夹
                                    if os.path.exists(img_dir):
                                        shutil.rmtree(img_dir, ignore_errors=True)
                                except Exception as e:
                                    print(f"处理目录出错 {img_dir}: {e}")
                                    continue
                        else:
                            print(f"跳过空的商品代码")
                            continue
                            
                    except (ValueError, TypeError) as e:
                        print(f"处理商品代码出错 '{product_code}': {e}")
                        continue
                            
                    self.tree.delete(item)
            
            # 更新产品列表
            print(f"要删除的商品代码列表: {product_codes_to_delete}")
            self.products = [p for p in self.products if int(float(str(p.get('Code', 0)))) not in product_codes_to_delete]
            self.filtered_products = [p for p in self.filtered_products if int(float(str(p.get('Code', 0)))) not in product_codes_to_delete]
            
            try:
                gc.collect()
                self.save_to_excel()
                self.refresh_tree()
                self.status_label.config(text="选中的商品已成功删除。")
            except Exception as e:
                print(f"保存更改时出错: {e}")
                self.status_label.config(text=f"保存更改时出错: {str(e)}")
                messagebox.showerror("错误", f"保存更改时出错: {str(e)}")
        else:
            self.status_label.config(text="未选择要删除的商品。")

    def clear_database(self):
        confirmation = simpledialog.askstring("Clear Database", "Type 'delete' to confirm:")
        if confirmation == 'delete':
            self.products.clear()
            self.filtered_products = self.products
            self.save_to_excel()
            self.refresh_tree()
        else:
            self.status_label.config(text="Database clearance cancelled.")

    def return_to_main_menu(self):
        try:
            print("Returning to main menu")
            self.is_running = False
            self.pasteboard_monitoring = False
            
            # 取消所有pending的after调用
            if hasattr(self, 'after_id'):
                try:
                    self.root.after_cancel(self.after_id)
                except Exception as e:
                    print(f"Error canceling after: {e}")
            
            if self.root:
                self.root.quit()
                self.root.destroy()
                
            # 显示主窗口
            import main
            main_app = main.ProductManager()
            main_app.root.deiconify()  # 显示主窗口
            
        except Exception as e:
            print(f"Error returning to main menu: {e}")

    def exit_program(self):
        print("Exiting program")
        self.is_running = False
        self.pasteboard_monitoring = False
        
        # 取消所有pending的after调用
        if hasattr(self, 'after_id'):
            try:
                self.root.after_cancel(self.after_id)
            except Exception as e:
                print(f"Error canceling after: {e}")
        
        if self.root:
            self.root.quit()
            self.root.destroy()

    def edit_selected_item(self, event):
        selected_items = self.tree.selection()
        if not selected_items:
            self.status_label.config(text="Please select an item to edit")
            return
        selected_item = selected_items[0]

    def on_double_click(self, event):
        # 获取选中的item
        item = self.tree.selection()[0]
        column = self.tree.identify_column(event.x)
        col_num = int(str(column).replace('#', ''))
        
        # 如果是StockQuantity列，不允许编辑
        if col_num == 7:  # StockQuantity列
            messagebox.showinfo("提示", "库存数量只能在库存管理界面修改")
            return
        
        # 获取当前值
        current_value = self.tree.item(item, "values")[col_num - 1]
        
        # 建编辑框
        def edit_cell(event=None):
            try:
                # 获取新值
                new_value = entry.get()
                
                # 处理特殊列的数据格式
                if col_num == 3:  # Cost列
                    try:
                        # 移除货币符号并转换为浮点数
                        new_value = float(str(new_value).replace('¥', '').replace('￥', '').strip())
                    except ValueError:
                        messagebox.showerror("错误", "请输入有效的价格")
                        return
                elif col_num == 7:  # StockQuantity列
                    try:
                        # 转换为整数
                        new_value = int(new_value)
                    except ValueError:
                        messagebox.showerror("错误", "请输入有效的整数")
                        return
                
                # 更新treeview
                values = list(self.tree.item(item, "values"))
                values[col_num - 1] = new_value
                self.tree.item(item, values=values)
                
                # 新Excel文件
                try:
                    wb = openpyxl.load_workbook('products.xlsx')
                    ws = wb.active
                    
                    # 找到应的行
                    code = values[4]  # Code在第5列
                    for row in range(2, ws.max_row + 1):
                        if str(ws.cell(row=row, column=5).value) == str(code):
                            # 更新单元格值
                            cell = ws.cell(row=row, column=col_num)
                            cell.value = new_value
                            
                            # 设置数字格式
                            if col_num == 3:  # Cost列
                                cell.number_format = '#,##0.00'
                            elif col_num == 7:  # StockQuantity列
                                cell.number_format = '0'
                            
                            break
                    
                    wb.save('products.xlsx')
                    print(f"Updated cell value in Excel: {new_value}")
                except Exception as e:
                    print(f"Error updating Excel: {e}")
                    messagebox.showerror("错误", f"更新Excel文件时出错: {str(e)}")
                
                entry.destroy()
                
            except Exception as e:
                print(f"Error in edit_cell: {e}")
                messagebox.showerror("错误", f"编辑单元格时出错: {str(e)}")
                entry.destroy()
        
        # 创建Entry件
        entry = ttk.Entry(self.tree)
        entry.insert(0, current_value)
        entry.select_range(0, tk.END)
        
        # 计算Entry的位置
        bbox = self.tree.bbox(item, column)
        if bbox:  # 确保bbox不为None
            entry.place(x=bbox[0], y=bbox[1], width=bbox[2])
        
        # 绑定事件
        entry.bind('<Return>', edit_cell)
        entry.bind('<Escape>', lambda e: entry.destroy())
        entry.bind('<FocusOut>', lambda e: entry.destroy())
        
        # 设置焦点
        entry.focus_set()

    def fix_missing_codes(self):
        """修复没有商品代码的产品"""
        try:
            fixed_count = 0
            for product in self.products:
                try:
                    # 检查商品代码是否缺失或无效
                    current_code = product.get("Code", 0)
                    if current_code == 0 or current_code is None or (isinstance(current_code, str) and not current_code.strip()):
                        # 生成新的商品代码
                        new_code = self.generate_unique_code()
                        if new_code:
                            product["Code"] = new_code
                            fixed_count += 1
                except Exception as e:
                    print(f"处理单个产品时出错: {e}")
                    continue

            if fixed_count > 0:
                # 保存更改
                self.save_to_excel()
                # 刷新显示
                self.refresh_tree()
                self.status_label.config(text=f"成功修复 {fixed_count} 个产品的商品代码")
            else:
                self.status_label.config(text="没有发现需要修复的商品代码")

        except Exception as e:
            print(f"修复商品代码时出错: {e}")
            self.status_label.config(text=f"修复商品代码时出错: {str(e)}")
            messagebox.showerror("错误", f"修复商品代码时出错: {str(e)}")

# Example use:
if __name__ == "__main__":
    products = []
    app = ProductManagerUI(products, None)