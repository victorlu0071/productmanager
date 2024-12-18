import tkinter as tk
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import requests
from datetime import datetime
import json
import threading
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import time
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import random
import tkinter.messagebox as messagebox

class ShoppingAssistantUI:
    def __init__(self):
        self.root = None
        self.browser = None
        self.recording = False
        self.purchases = []
        self.product_cache = {}  # 用于存储产品页面的图片URL
        self.cached_urls = set()  # 添加URL缓存集合
        self.current_url = None  # 添加当前URL记录
        self.processed_products = set()  # 添加已处理商品记录
        self.last_product_id = None  # 添加最后访问的产品ID记录
        self.setup_styles()  # 设置样式
        self.init_gui()
        
    def setup_styles(self):
        """设置UI样式"""
        # 定义颜色
        self.colors = {
            'primary': '#2196F3',  # 主色调
            'secondary': '#64B5F6',  # 次要色调
            'success': '#4CAF50',  # 成功色
            'warning': '#FFC107',  # 警告色
            'error': '#F44336',  # 错误色
            'background': '#F5F5F5',  # 背景色
            'text': '#212121',  # 文本色
            'light_text': '#757575'  # 浅色文本
        }

        # 创建自定义样式
        style = ttk.Style()
        style.theme_use('clam')

        # 配置Treeview样式
        style.configure("Treeview",
            background=self.colors['background'],
            foreground=self.colors['text'],
            fieldbackground=self.colors['background'],
            rowheight=30,
            font=('Microsoft YaHei UI', 10))
        
        style.configure("Treeview.Heading",
            background=self.colors['primary'],
            foreground="white",
            relief="flat",
            font=('Microsoft YaHei UI', 10, 'bold'))
        
        style.map("Treeview.Heading",
            background=[('active', self.colors['secondary'])])

        # 配置按钮样式
        style.configure("Primary.TButton",
            background=self.colors['primary'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10))
        
        style.map("Primary.TButton",
            background=[('active', self.colors['secondary'])])

        # 配置危险按钮样式
        style.configure("Danger.TButton",
            background=self.colors['error'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10))
        
        style.map("Danger.TButton",
            background=[('active', '#E57373')])

        # 配置标签样式
        style.configure("Info.TLabel",
            background=self.colors['background'],
            foreground=self.colors['text'],
            font=('Microsoft YaHei UI', 10))

        # 配置输入框样式
        style.configure("Custom.TEntry",
            fieldbackground="white",
            padding=(5, 5))

    def init_gui(self):
        try:
            self.root = tk.Tk()
            self.root.title("采购助手")
            
            self.root.protocol("WM_DELETE_WINDOW", self.return_to_main_menu)
            
            # 设置窗口样式
            self.root.configure(bg=self.colors['background'])
            
            window_width = 1080
            window_height = 700
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            position_top = int((screen_height - window_height) / 2)
            position_left = int((screen_width - window_width) / 2)
            self.root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

            # 创建主框架
            main_frame = ttk.Frame(self.root)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

            # 控制按钮
            control_frame = ttk.Frame(main_frame)
            control_frame.pack(pady=10)

            self.start_button = ttk.Button(control_frame, 
                                         text="启动浏览器", 
                                         command=self.start_recording,
                                         style="Primary.TButton")
            self.start_button.pack(side=tk.LEFT, padx=5)

            self.continue_button = ttk.Button(control_frame, 
                                            text="继续记录", 
                                            command=self.continue_recording,
                                            style="Primary.TButton")
            self.continue_button.pack(side=tk.LEFT, padx=5)
            self.continue_button.config(state=tk.DISABLED)

            self.stop_button = ttk.Button(control_frame, 
                                        text="停止记录", 
                                        command=self.stop_recording,
                                        style="Danger.TButton")
            self.stop_button.pack(side=tk.LEFT, padx=5)
            self.stop_button.config(state=tk.DISABLED)

            # Treeview区域
            tree_frame = ttk.Frame(main_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)

            # 创建Treeview的滚动条
            tree_scroll = ttk.Scrollbar(tree_frame)
            tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

            # Treeview
            self.tree = ttk.Treeview(tree_frame, 
                                   columns=("Time", "Website", "Product", "Price"),
                                   show='headings',
                                   style="Treeview",
                                   yscrollcommand=tree_scroll.set)
            
            tree_scroll.config(command=self.tree.yview)

            # 设置列宽和表头
            column_headers = {
                "Time": "时间",
                "Website": "网站",
                "Product": "商品",
                "Price": "价格"
            }

            for col in column_headers:
                self.tree.heading(col, text=column_headers[col])
                self.tree.column(col, width=200, anchor='center')

            self.tree.pack(fill=tk.BOTH, expand=True)

            # 按钮区域
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(pady=10)

            back_button = ttk.Button(button_frame, 
                                   text="返回主菜单", 
                                   command=self.return_to_main_menu,
                                   style="Primary.TButton")
            back_button.pack(pady=5)

            # 状态标签
            self.status_label = ttk.Label(main_frame, 
                                        text="就绪", 
                                        anchor="w",
                                        style="Info.TLabel")
            self.status_label.pack(fill=tk.X, padx=10, pady=5)

            self.root.mainloop()

        except Exception as e:
            print(f"Error initializing GUI: {e}")
            self.return_to_main_menu()

    def start_recording(self):
        try:
            # 设置Chrome选项
            chrome_options = webdriver.ChromeOptions()
            
            # 设置用户数据目录
            user_data_dir = os.path.join(os.getcwd(), "chrome_user_data")
            os.makedirs(user_data_dir, exist_ok=True)
            chrome_options.add_argument(f"user-data-dir={user_data_dir}")
            
            # 添加其他选项
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-features=TranslateUI")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-popup-blocking")
            
            # 添加WebGL相关的选项
            chrome_options.add_argument("--enable-unsafe-swiftshader")
            chrome_options.add_argument("--ignore-gpu-blocklist")
            chrome_options.add_argument("--enable-gpu-rasterization")
            
            # 添加控制新标签页行为的参数
            chrome_options.add_argument("--disable-features=IsolateOrigins,site-per-process")
            chrome_options.add_argument("--disable-site-isolation-trials")
            
            # 添加实验性选项
            chrome_options.add_experimental_option("prefs", {
                "profile.default_content_settings.popups": 0,
                "profile.default_content_setting_values.automatic_downloads": 1,
                "profile.default_content_settings.notifications": 2,
                "profile.default_content_setting_values.new_window_disposition": 1,  # 强制在当前标签页打开
                "profile.default_content_setting_values.new_tab_disposition": 1,     # 强制在当前标签页打开
                "profile.default_content_setting_values.window_open_disposition": 1,  # 强制在当前标签页打开
                "profile.default_content_setting_values.tabs": 2
            })
            
            # 初始化浏览器
            service = Service(ChromeDriverManager().install())
            self.browser = webdriver.Chrome(service=service, options=chrome_options)
            
            # 首先访问登录页面
            print("Accessing login page...")
            self.browser.get("https://login.1688.com/member/signin.htm")
            
            # 更新状态标签
            self.status_label.config(text="请在浏览器中登录，然后点击 'Continue Recording' 按钮...")
            
            # 禁用开始按钮，启用继续按钮
            self.start_button.config(state=tk.DISABLED)
            self.continue_button.config(state=tk.NORMAL)
            
        except Exception as e:
            self.status_label.config(text=f"Error starting browser: {e}")
            print(f"Error details: {str(e)}")
            # 如果出错，重新启用开始按钮
            self.start_button.config(state=tk.NORMAL)

    def continue_recording(self):
        try:
            # 确认登录状态
            time.sleep(2)
            
            self.recording = True
            self.continue_button.config(state=tk.DISABLED)
            self.stop_button.config(state=tk.NORMAL)
            self.status_label.config(text="Recording started...")
            
            # 在新线程中开始监控
            threading.Thread(target=self.monitor_pages, daemon=True).start()
            
        except Exception as e:
            self.status_label.config(text=f"Error starting recording: {e}")

    def stop_recording(self):
        try:
            self.recording = False
            if self.browser:
                try:
                    self.browser.quit()
                except:
                    pass  # 忽略关闭览器时的错误
            
            # 清理缓存
            self.clear_cache()
            
            self.browser = None
            self.current_url = None  # 重置当前URL
            self.processed_products.clear()  # 重置处理状态
            self.start_button.config(state=tk.NORMAL)
            self.continue_button.config(state=tk.DISABLED)
            self.stop_button.config(state=tk.DISABLED)
            self.status_label.config(text="Recording stopped")
            self.save_purchases()
        except Exception as e:
            print(f"Error stopping recording: {e}")

    def clear_cache(self):
        """清理所有缓存文件和数据"""
        try:
            # 清理临时图片目录
            temp_dir = "temp_images"
            if os.path.exists(temp_dir):
                import shutil
                shutil.rmtree(temp_dir)
                print("Cleared temporary images directory")
            
            # 重置缓存数据
            self.product_cache.clear()
            self.cached_urls.clear()
            self.last_product_id = None
            print("Cleared cache data")
            
        except Exception as e:
            print(f"Error clearing cache: {e}")

    def monitor_pages(self):
        """监控浏览器标签页，当前标签关闭时自动切换到下一个标签"""
        last_processed_url = None
        
        while self.recording:
            try:
                # 检查浏览器是否还在运行
                if not self.browser.window_handles:
                    print("Browser closed. Stopping recording...")
                    self.root.after(0, self.stop_recording)
                    break

                # 检查当前标签页是否还存在
                try:
                    current_window = self.browser.current_window_handle
                except:
                    # 如果当前标签页不存在，切换到可用的标签页
                    if self.browser.window_handles:
                        print("Current tab closed, switching to next available tab...")
                        # 获取所有标签页
                        all_handles = self.browser.window_handles
                        # 切换到第一个可用标签页
                        self.browser.switch_to.window(all_handles[0])
                        # 重置处理状态
                        last_processed_url = None
                        time.sleep(0.5)  # 等待切换完成
                        continue
                    else:
                        print("No available tabs. Stopping recording...")
                        self.root.after(0, self.stop_recording)
                        break

                # 获取当前标签页URL
                try:
                    current_url = self.browser.current_url
                    # 如果URL没有变化，跳过处理
                    if current_url == last_processed_url:
                        time.sleep(0.3)
                        continue
                except:
                    # 如果获取URL失败，可能是标签页正在加载或已关闭
                    print("Failed to get current URL, checking tab status...")
                    time.sleep(0.5)
                    continue

                # 处理不同类型的页面
                if "detail.1688.com/offer" in current_url:
                    # 产品详情页面
                    product_id = self.extract_product_id(current_url)
                    if product_id and product_id != self.last_product_id:
                        print(f"New product detected: {product_id}")
                        self.last_product_id = product_id
                        if current_url not in self.cached_urls:
                            self.cache_product_images()
                    last_processed_url = current_url

                elif "order.1688.com/order/smart_make_order.htm" in current_url:
                    # 订单页面
                    if current_url != last_processed_url:  # 只在URL改变时处理
                        print("Found order page, recording product info...")
                        if self.last_product_id:
                            self.record_product_info()
                        else:
                            print("No product ID found in cache")
                    last_processed_url = current_url

                elif "trade.1688.com/order/trade_flow.htm" in current_url:
                    # 收银台页面
                    if "收银台" in self.browser.page_source and current_url != last_processed_url:
                        print("Found checkout page, saving to database...")
                        if self.last_product_id:
                            # 遍历所有规格的缓存记录
                            saved_count = 0
                            for cache_key in list(self.product_cache.keys()):
                                if cache_key.startswith(self.last_product_id) and '_' in cache_key:  # 确保是规格记录而不是图片缓存
                                    try:
                                        if self.save_to_database(cache_key):
                                            saved_count += 1
                                            print(f"Successfully saved specification {cache_key} to database")
                                    except Exception as e:
                                        print(f"Error saving specification {cache_key}: {e}")
                            
                            if saved_count == 0:
                                print("No specifications were saved to database")
                            else:
                                print(f"Successfully saved {saved_count} specifications to database")
                        last_processed_url = current_url

                time.sleep(0.3)  # 低CPU使用率

            except Exception as e:
                print(f"Error in monitor_pages: {e}")
                import traceback
                print(traceback.format_exc())
                time.sleep(0.5)

    def record_product_info(self):
        """记录产品信息，处理所有选择的规格"""
        try:
            # 等待页面加载
            WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "offer-link"))
            )

            # 获取产品名
            product_name = self.browser.find_element(By.CLASS_NAME, "offer-link").text.strip()
            print(f"Recording info for product: {product_name}")
            
            # 获取所有cargo-inner元素
            cargo_inners = self.browser.find_elements(By.CLASS_NAME, "cargo-inner")
            print(f"Found {len(cargo_inners)} different specifications")
            
            if self.last_product_id:
                # 获取原图片信息
                original_cache = self.product_cache.get(self.last_product_id, {})
                image_urls = original_cache.get('images', [])
                
                if image_urls:
                    # 遍历每个规格
                    for cargo_inner in cargo_inners:
                        try:
                            # 获取价格
                            price_element = cargo_inner.find_element(By.CLASS_NAME, "cargo-unit-price")
                            price = price_element.text.strip()
                            
                            # 获取规格信息
                            cargo_info = cargo_inner.find_element(By.CLASS_NAME, "cargo-info")
                            cargo_desc = cargo_info.find_element(By.CLASS_NAME, "cargo-desc")
                            cargo_spec = cargo_desc.find_element(By.CLASS_NAME, "cargo-spec")
                            spec_items = cargo_spec.find_elements(By.CLASS_NAME, "spec-item")
                            
                            # 收集规格信息
                            current_specs = []
                            for item in spec_items:
                                spec_text = item.text.strip()
                                if spec_text:
                                    spec_text = spec_text.rstrip(";")
                                    current_specs.append(spec_text)
                            
                            # 创建规格组合
                            spec_combination = ";".join(current_specs) if current_specs else "默认格:默认"
                            
                            # 创建缓存键
                            cache_key = f"{self.last_product_id}_{spec_combination}"
                            
                            # 获取数量信息
                            try:
                                quantity_element = cargo_inner.find_element(By.CLASS_NAME, "amount-input")
                                quantity = quantity_element.get_attribute("value")
                            except:
                                quantity = "1"
                                print("Could not find quantity, using default value: 1")
                            
                            if cache_key not in self.product_cache:
                                product_code = self.generate_unique_code()
                                
                                # 创建图片目录
                                spec_image_dir = os.path.join("product_images", str(product_code))
                                os.makedirs(spec_image_dir, exist_ok=True)
                                
                                # 复制图片
                                temp_dir = os.path.join("temp_images", self.last_product_id)
                                copied_images = []
                                for i, url in enumerate(image_urls):
                                    temp_path = os.path.join(temp_dir, f"image_{i+1}.jpg")
                                    if os.path.exists(temp_path):
                                        spec_path = os.path.join(spec_image_dir, f"image_{i+1}.jpg")
                                        import shutil
                                        shutil.copy2(temp_path, spec_path)
                                        copied_images.append(spec_path)
                                
                                self.product_cache[cache_key] = {
                                    'name': product_name,
                                    'specs': spec_combination,
                                    'price': price,
                                    'code': product_code,
                                    'quantity': quantity,
                                    'images': image_urls
                                }
                                print(f"Recorded product info: {product_name}, {spec_combination}, {price}, Quantity: {quantity}, Code: {product_code}")
                            
                        except Exception as e:
                            print(f"Error processing specification: {e}")
                            continue
                else:
                    print("No images found in cache")
            else:
                print("No product ID found in cache")
            
        except Exception as e:
            print(f"Error recording product info: {e}")
            import traceback
            print(traceback.format_exc())

    def save_to_database(self, cache_key):
        """保存指定缓存键的产品信息到数据库"""
        try:
            if not cache_key or cache_key not in self.product_cache:
                print(f"No product information to save for cache key: {cache_key}")
                return False
            
            cache_data = self.product_cache[cache_key]
            required_keys = ['name', 'specs', 'price', 'images', 'code', 'quantity']
            
            # 检查必需的键
            missing_keys = [key for key in required_keys if key not in cache_data]
            if missing_keys:
                print(f"Missing required keys in cache for {cache_key}: {missing_keys}")
                return False
            
            # 从缓存键中提取商品ID（修改这部分）
            try:
                product_id = cache_key.split('_')[0]  # 直接获取下划线前的部分作为商品ID
                if not product_id or product_id == 'None':
                    print(f"Invalid product ID in cache key: {cache_key}")
                    return False
                product_link = f"https://detail.1688.com/offer/{product_id}.html"
                print(f"Generated product link: {product_link}")  # 添加日志
            except Exception as e:
                print(f"Error extracting product ID from cache key: {e}")
                return False
            
            # 添加购买记录
            self.add_purchase_record(
                product_link,  # 使用商品详情页链接
                f"{cache_data['name']} [{cache_data['specs']}]",
                cache_data['price'],
                cache_data['quantity'],
                cache_data['images'],
                cache_data['code']
            )
            
            print(f"Successfully saved product {cache_data['code']} to database")
            return True
            
        except Exception as e:
            print(f"Error saving to database: {e}")
            import traceback
            print(traceback.format_exc())
            return False

    def extract_product_info(self, driver):
        try:
            # 等待价格元素加载
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "span.amount"))
            )
            
            # 获取价格
            price_element = driver.find_element(By.CSS_SELECTOR, "span.amount")
            price = price_element.text.strip()
            if not price:
                price = "0.00"
            
            # 移除货币符号并转换为浮点数
            try:
                price = float(price.replace('¥', '').replace('￥', '').strip())
            except ValueError:
                price = 0.00
            
            # 获取规格信息
            specs = []
            try:
                spec_elements = driver.find_elements(By.CSS_SELECTOR, ".sku-name")
                for spec in spec_elements:
                    spec_text = spec.text.strip()
                    if spec_text:
                        specs.append(spec_text)
            except Exception as e:
                print(f"Error getting specifications: {e}")
            
            # 获取商品名
            try:
                name_element = driver.find_element(By.CSS_SELECTOR, ".product-title")
                name = name_element.text.strip()
            except Exception as e:
                print(f"Error getting product name: {e}")
                name = ""
            
            # 获取商品链接
            product_url = driver.current_url
            
            # 生成唯一的商品代码
            product_code = str(random.randint(1000000, 9999999))
            
            # 创建商品信息字典
            product_info = {
                "Name": name,
                "Specs": " | ".join(specs) if specs else "",
                "Cost": f"{price:.2f}",
                "Link": product_url,
                "Code": product_code,
                "LocationCode": "",
                "StockQuantity": "0",
                "AddedDate": datetime.now().strftime("%Y-%m-%d")
            }
            
            return product_info
            
        except Exception as e:
            print(f"Error extracting product info: {e}")
            return None

    def extract_product_id(self, url):
        # 从URL中提取产品ID
        try:
            return url.split('offer/')[1].split('.html')[0]
        except:
            return None

    def generate_unique_code(self):
        while True:
            new_code = random.randint(1000000, 9999999)  # 直接生成整数
            try:
                # 检查Excel文件中是否存在此Code
                wb = openpyxl.load_workbook('products.xlsx')
                ws = wb.active
                codes = [int(str(ws.cell(row=row, column=5).value)) for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=5).value is not None]  # 获取现有的代码转换为整数
                if new_code not in codes:
                    wb.close()
                    return new_code  # 返回整数
            except FileNotFoundError:
                return new_code  # 如果文件不存在，直接返回新生成的整数
            except Exception as e:
                print(f"Error checking unique code: {e}")
                wb.close()
                continue

    def download_image_with_retry(self, url, save_path, max_retries=3):
        """下载图片并在失败时重试"""
        for attempt in range(max_retries):
            try:
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    img_data = io.BytesIO(response.content)
                    with Image.open(img_data) as img:
                        width, height = img.size
                        if width >= 100 and height >= 100:
                            with open(save_path, 'wb') as f:
                                f.write(response.content)
                            return True
                        else:
                            print(f"Image too small: {width}x{height}")
                            return False
            except Exception as e:
                print(f"Attempt {attempt + 1} failed for {url}: {e}")
                if attempt == max_retries - 1:  # 如果是最后一次尝试
                    if isinstance(e, requests.exceptions.SSLError):
                        print("SSL Error occurred, refreshing page and trying one last time...")
                        try:
                            # 刷新当前页面
                            self.browser.refresh()
                            time.sleep(2)  # 等待页面加载
                            # 最后一次尝试
                            response = requests.get(url, timeout=10)
                            if response.status_code == 200:
                                img_data = io.BytesIO(response.content)
                                with Image.open(img_data) as img:
                                    width, height = img.size
                                    if width >= 100 and height >= 100:
                                        with open(save_path, 'wb') as f:
                                            f.write(response.content)
                                        return True
                        except Exception as final_e:
                            print(f"Final attempt after refresh failed: {final_e}")
                            return False
                time.sleep(1)  # 在重试之间等待1秒
        return False

    def cache_product_images(self):
        """缓存当前标签页中的产品图片"""
        try:
            current_url = self.browser.current_url
            if current_url in self.cached_urls:  # 检查URL是否已经缓存过
                return
            
            product_id = self.extract_product_id(current_url)
            if not product_id:
                return

            print(f"Caching images for product {product_id} in current tab")
            
            # 创建队列用于存储找到的图片URL
            from queue import Queue
            image_queue = Queue()
            
            # 创建事件用于控制线程终止
            found_images = threading.Event()
            
            def method1():
                """方法1：获取od-gallery-turn-item-wrapper中的图片"""
                try:
                    if found_images.is_set():  # 如果另一个方法已经找到图片，直接返回
                        return
                    
                    WebDriverWait(self.browser, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "od-gallery-turn-item-wrapper"))
                    )
                    image_containers = self.browser.find_elements(By.CLASS_NAME, "od-gallery-turn-item-wrapper")
                    if image_containers:
                        for container in image_containers[:6]:
                            if found_images.is_set():  # 检查是否需要终止
                                break
                            img_element = container.find_element(By.CLASS_NAME, "od-gallery-img")
                            img_url = img_element.get_attribute("src")
                            if img_url and self.check_image_size(img_url):
                                image_queue.put(img_url)
                        if image_queue.qsize() > 0:
                            found_images.set()  # 设置事件，通知另一个线程
                            print("Method 1 found images")
                except Exception as e:
                    print(f"Method 1 failed: {e}")

            def method2():
                """方法2：获取detail-gallery-turn-wrapper中的图片"""
                try:
                    if found_images.is_set():  # 如果另一个方法已经找到图片，直接返回
                        return
                    
                    WebDriverWait(self.browser, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "detail-gallery-turn-wrapper"))
                    )
                    
                    image_containers = self.browser.find_elements(By.CLASS_NAME, "detail-gallery-turn-wrapper")
                    print(f"Found {len(image_containers)} detail-gallery-turn-wrapper elements")
                    
                    for container in image_containers[:6]:
                        if found_images.is_set():  # 检查是否需要终止
                            break
                        try:
                            img_element = container.find_element(By.TAG_NAME, "img")
                            img_url = img_element.get_attribute("src")
                            if img_url and self.check_image_size(img_url):
                                image_queue.put(img_url)
                                print(f"Found image URL: {img_url}")
                        except Exception as e:
                            print(f"Error processing image container: {e}")
                            continue
                    
                    if image_queue.qsize() > 0:
                        found_images.set()  # 设置事件，通知另一个线程
                        print("Method 2 found images")
                    
                except Exception as e:
                    print(f"Method 2 failed: {e}")

            # 创建并启动两个线程
            thread1 = threading.Thread(target=method1, daemon=True)
            thread2 = threading.Thread(target=method2, daemon=True)
            
            thread1.start()
            thread2.start()
            
            # 等待任一线程找到图片或超时
            found_images.wait(timeout=10)
            
            # 从队列中获取所有图片URL
            image_urls = []
            seen_urls = set()  # 用于去重
            while not image_queue.empty():
                url = image_queue.get()
                if url not in seen_urls:
                    image_urls.append(url)
                    seen_urls.add(url)
                    if len(image_urls) >= 6:  # 最多取6张图片
                        break

            if not image_urls:
                print("No valid images found in current tab")
                return

            print(f"Found {len(image_urls)} valid images")
            
            # 缓存图片信息
            self.product_cache[product_id] = {
                'images': image_urls
            }
            self.cached_urls.add(current_url)
            
            # 立即下载图片到临时目录
            temp_dir = os.path.join("temp_images", product_id)
            os.makedirs(temp_dir, exist_ok=True)
            
            # 使用线程池并行下载图片
            from concurrent.futures import ThreadPoolExecutor
            
            def download_image(args):
                i, url = args
                img_path = os.path.join(temp_dir, f"image_{i+1}.jpg")
                if self.download_image_with_retry(url, img_path):
                    print(f"Successfully downloaded image {i+1} for product {product_id}")
                    return url
                print(f"Failed to download image {i+1} after all retries")
                return None
            
            with ThreadPoolExecutor(max_workers=3) as executor:
                successful_downloads = list(filter(None, executor.map(
                    download_image, 
                    enumerate(image_urls)
                )))

            # 更新缓存中的图片列表，只保留成功下载的图片URL
            if successful_downloads:
                self.product_cache[product_id]['images'] = successful_downloads
            else:
                print("No images were successfully downloaded")
                self.product_cache.pop(product_id, None)
                self.cached_urls.remove(current_url)

        except Exception as e:
            print(f"Error caching product images: {e}")
            import traceback
            print(traceback.format_exc())

    def check_image_size(self, url):
        """检查图片尺寸否符合要求"""
        try:
            response = requests.get(url)
            if response.status_code == 200:
                img_data = io.BytesIO(response.content)
                with Image.open(img_data) as img:
                    width, height = img.size
                    if width >= 100 and height >= 100:
                        print(f"Image size check passed: {width}x{height}")
                        return True
                    else:
                        print(f"Image too small: {width}x{height}")
                        return False
        except Exception as e:
            print(f"Error checking image size: {e}")
            return False
        return False

    def extract_1688_info(self):
        try:
            # 加确的等待时间
            time.sleep(2)
            
            # 等待页面全加载
            WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            # 取当前页面URL作为链接
            current_url = self.browser.current_url

            # 获取产品链接和名称
            try:
                # 在收银台页面，产品同的位置
                if "trade.1688.com/order/trade_flow.htm" in current_url:
                    # 等待产品信息加载
                    WebDriverWait(self.browser, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "order-item"))
                    )
                    # 获取产品名称
                    product_name = self.browser.find_element(By.CLASS_NAME, "order-item-title").text.strip()
                    # 获取产品链接
                    product_link = self.browser.find_element(By.CLASS_NAME, "order-item-title").get_attribute("href")
                    product_id = self.last_product_id  # 使用之前缓存的产品ID
                else:
                    # 原有的产品页面逻辑
                    product_link = WebDriverWait(self.browser, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "offer-link"))
                    )
                    product_id = self.extract_product_id(product_link.get_attribute("href"))
                    product_name = product_link.text

                print(f"Found product: {product_name}")
            except Exception as e:
                print(f"Error getting product info: {e}")
                return

            # 获取价格
            try:
                if "trade.1688.com/order/trade_flow.htm" in current_url:
                    # 在收银台页面获取价格
                    price_element = self.browser.find_element(By.CLASS_NAME, "order-item-price")
                    price = price_element.text.strip()
                else:
                    # 原有的产品页面价格获取逻辑
                    price = WebDriverWait(self.browser, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "amount"))
                    ).text
                print(f"Found price: {price}")
            except Exception as e:
                print(f"Error getting price: {e}")
                price = "N/A"

            # 获取规格信息
            try:
                if "trade.1688.com/order/trade_flow.htm" in current_url:
                    # 在收银台页面取规格
                    spec_items = self.browser.find_elements(By.CLASS_NAME, "order-item-sku")
                    specs = []
                    for item in spec_items:
                        spec_text = item.text.strip()
                        if spec_text:
                            specs.append(spec_text)
                else:
                    # 原有产品页面规格获取逻辑
                    spec_items = self.browser.find_elements(By.CLASS_NAME, "spec-item")
                    specs = []
                    for item in spec_items:
                        spec_text = item.text.strip()
                        if spec_text:
                            specs.append(spec_text)
                specs_str = "; ".join(specs)
                print(f"Found specifications: {specs_str}")
            except Exception as e:
                print(f"Error getting specifications: {e}")
                specs_str = ""

            # 检查是否已经处理过该商品
            if product_id in self.processed_products:
                print(f"Product {product_id} already processed, skipping...")
                return

            # 检查是否有缓存的图片
            if product_id in self.product_cache:
                try:
                    cache_data = self.product_cache[product_id]
                    image_urls = cache_data['images']
                    product_code = cache_data['code']  # 使用缓存中的Code
                    
                    # 在这里取产品名称和价格
                    cache_data['name'] = product_name
                    cache_data['price'] = price
                    cache_data['specs'] = specs_str  # 添加规格信息
                    cache_data['timestamp'] = datetime.now()
                    
                    # 添加购买记录，使用完整URL为链接
                    # 将产品名称和规格组合在一起
                    full_product_name = f"{product_name} [{specs_str}]" if specs_str else product_name
                    self.add_purchase_record(current_url, full_product_name, price, "1", image_urls, product_code)
                    
                    # 标记为处理
                    self.processed_products.add(product_id)
                    
                    print(f"Successfully recorded purchase: {full_product_name} with code {product_code}")
                except Exception as e:
                    print(f"Error processing cached images: {e}")
            else:
                print(f"No cached images found for product {product_id}")

        except Exception as e:
            print(f"Error in extract_1688_info: {e}")
            import traceback
            print(traceback.format_exc())

    def extract_taobao_info(self):
        # 淘宝特定提取逻辑
        try:
            product_name = WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "tb-main-title"))
            ).text
            price = self.browser.find_element(By.CLASS_NAME, "tb-rmb-num").text
            # 获取图片URL
            image_elements = self.browser.find_elements(By.CSS_SELECTOR, ".tb-thumb img")
            image_urls = [img.get_attribute("src") for img in image_elements]
            
            self.download_images(image_urls, product_name)
            self.add_purchase_record("taobao.com", product_name, price, "1", image_urls)
            
        except Exception as e:
            print(f"Error extracting Taobao info: {e}")

    def download_images(self, image_urls, product_code):
        # 创建图片保存目录
        save_dir = os.path.join("product_images", product_code)
        os.makedirs(save_dir, exist_ok=True)

        for i, url in enumerate(image_urls):
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    file_path = os.path.join(save_dir, f"image_{i+1}.jpg")
                    with open(file_path, 'wb') as f:
                        f.write(response.content)
                    print(f"Saved image {i+1} for product {product_code}")
            except Exception as e:
                print(f"Error downloading image {url}: {e}")

    def add_purchase_record(self, website, product, price, quantity, image_urls, product_code):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 分离产品名称和规格
        if "[" in product and "]" in product:
            product_name = product.split("[")[0].strip()
            specs = product.split("[")[1].rstrip("]").strip()
        else:
            product_name = product
            specs = ""
        
        # 处理价格格式
        try:
            # 移除所有非数字和小数点的字符
            price_str = ''.join(char for char in str(price) if char.isdigit() or char == '.')
            
            # 如果字符串为空或只包含小数点，设置为0
            if not price_str or price_str == '.':
                price_float = 0.0
            else:
                # 如果有多个小数点，只保留第一个
                if price_str.count('.') > 1:
                    parts = price_str.split('.')
                    price_str = parts[0] + '.' + ''.join(parts[1:])
                
                # 转换为浮点数
                price_float = float(price_str)
                
            # 格式化价格显示（用于UI显示）
            formatted_price = f"¥{price_float:.2f}"
            # 数据库中存储数字
            db_price = price_float
        except:
            formatted_price = "¥0.00"
            db_price = 0.0
            
        # 处理数量格式
        try:
            quantity_int = int(str(quantity).strip())
        except:
            quantity_int = 0
            
        # 确保product_code是整数
        try:
            product_code = int(str(product_code).strip())
        except:
            product_code = 0
            
        record = {
            "time": timestamp,
            "website": website,
            "product": product_name,
            "specs": specs,
            "price": formatted_price,
            "product_code": product_code  # 存储为整数
        }
        self.purchases.append(record)
        
        # 更新UI
        self.root.after(0, lambda: self.tree.insert("", 0, values=(
            timestamp, website, product, formatted_price
        )))

        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                # 准备新的商品数据
                image_paths = []
                save_dir = os.path.join("product_images", str(product_code))  # 转换为字符串用于路径
                os.makedirs(save_dir, exist_ok=True)
                
                # 下载并保存原始图片
                saved_images = []
                temp_files = []  # 用于跟踪临时文件
                
                for i, url in enumerate(image_urls):
                    if i >= 6:  # 最多保存6张图片
                        break
                    try:
                        response = requests.get(url)
                        if response.status_code == 200:
                            # 保存原始图片
                            img_path = os.path.join(save_dir, f"image_{i+1}.jpg")
                            with open(img_path, 'wb') as f:
                                f.write(response.content)
                            image_paths.append(img_path)
                            saved_images.append(response.content)
                    except Exception as e:
                        print(f"Error downloading image {url}: {e}")

                # 创建新的商品记录
                new_product = {
                    "Name": product_name,
                    "Specs": specs,
                    "Cost": db_price,  # 存储数字格式的价格
                    "Link": website,
                    "Code": product_code,  # 存储为整数
                    "LocationCode": "",
                    "StockQuantity": quantity_int,  # 存储数字格式的数量
                    "AddedDate": datetime.now().strftime("%Y-%m-%d")
                }

                # 读取或创建Excel文件
                try:
                    wb = openpyxl.load_workbook('products.xlsx')
                    ws = wb.active
                except FileNotFoundError:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    # 添加表头
                    headers = ["Name", "Specs", "Cost", "Link", "Code", "LocationCode", "StockQuantity", "AddedDate"]
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    # 设置数字列的格式
                    cost_col = ws.column_dimensions['C']
                    code_col = ws.column_dimensions['E']  # Code列
                    stock_col = ws.column_dimensions['G']
                    cost_col.number_format = '#,##0.00'
                    code_col.number_format = '0'  # 整数格式
                    stock_col.number_format = '0'
                except PermissionError:
                    retry_count += 1
                    if retry_count >= max_retries:
                        self.status_label.config(text="无法访问Excel文件，请确保文件未被其他程序打开")
                        messagebox.showerror("错误", "请关闭Excel文件后重试")
                        raise
                    time.sleep(1)
                    continue

                # 查找最后一行
                last_row = ws.max_row
                product_row = None

                # 查找是否已存在相同Code的商品
                for row in range(2, last_row + 1):
                    cell_value = ws.cell(row=row, column=5).value  # Code在第5列
                    try:
                        if int(str(cell_value)) == product_code:  # 确保比时都是整数
                            product_row = row
                            break
                    except (ValueError, TypeError):
                        continue

                if not product_row:
                    product_row = last_row + 1

                # 写入商品数据
                for col, (key, value) in enumerate(new_product.items(), 1):
                    cell = ws.cell(row=product_row, column=col, value=value)
                    # 为价格、代码和数量列设置数字格式
                    if col == 3:  # Cost列
                        cell.number_format = '#,##0.00'
                    elif col == 5:  # Code列
                        cell.number_format = '0'
                    elif col == 7:  # StockQuantity
                        cell.number_format = '0'

                # 调整行以适应图片
                ws.row_dimensions[product_row].height = 150  # 增加行以适应多图片

                # 添加图片到Excel（略图）
                for i, img_data in enumerate(saved_images):
                    if i >= 6:  # 最多显示6张图片
                        break
                    try:
                        # 将图片数据转换为PIL Image
                        img = Image.open(io.BytesIO(img_data))
                        # 调整片大小用于Excel显示
                        img.thumbnail((100, 100))
                        # 保存调整后的图片临时文件
                        temp_path = os.path.join(save_dir, f"temp_img_{product_row}_{i}.png")
                        img.save(temp_path)
                        temp_files.append(temp_path)
                        # 创建Excel图片对
                        xl_img = XLImage(temp_path)
                        # 计算图片位置（从第9列开始放置图片，因为添加了Specs）
                        col_letter = openpyxl.utils.get_column_letter(9 + i)
                        # 添加片到单元格
                        ws.add_image(xl_img, f"{col_letter}{product_row}")
                    except Exception as e:
                        print(f"Error adding image {i+1} to Excel: {e}")

                try:
                    # 保存Excel文件
                    wb.save('products.xlsx')
                    wb.close()  # 确保闭工簿
                    break  # 如果成功保存，跳出重试循环
                except PermissionError:
                    retry_count += 1
                    if retry_count >= max_retries:
                        self.status_label.config(text="无法保存Excel文件，请确保文件未被其他程序打开")
                        messagebox.showerror("��误", "请关闭Excel文件后重试")
                        raise
                    time.sleep(1)  # 等待1秒后重试
                    continue
                finally:
                    # 清理临时文件
                    for temp_file in temp_files:
                        try:
                            if os.path.exists(temp_file):
                                os.remove(temp_file)
                        except Exception as e:
                            print(f"Error deleting temp file {temp_file}: {e}")

            except Exception as e:
                print(f"Error saving to products.xlsx: {e}")
                import traceback
                print(traceback.format_exc())
                retry_count += 1
                if retry_count >= max_retries:
                    raise
                time.sleep(1)  # 等待1秒后重试

    def save_purchases(self):
        try:
            with open("purchases.json", "w", encoding="utf-8") as f:
                json.dump(self.purchases, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving purchases: {e}")

    def return_to_main_menu(self):
        self.stop_recording()
        if self.root:
            self.root.quit()
            self.root.destroy()
        import main
        main.ProductManager()

    def inject_js_code(self):
        """注入JavaScript代码以制链接行为"""
        try:
            self.browser.execute_script("""
                // 覆盖window.open方法
                window.open = function(url, target, features) {
                    if (url) {
                        window.location.href = url;
                        return window;
                    }
                    return null;
                };
                
                // 修改所有链接的target属性
                function modifyLinks() {
                    document.querySelectorAll('a').forEach(function(link) {
                        link.target = '_self';
                        link.addEventListener('click', function(e) {
                            if (this.href && !this.href.startsWith('javascript:')) {
                                e.preventDefault();
                                window.location.href = this.href;
                            }
                        });
                    });
                }
                
                // 立即执行一次
                modifyLinks();
                
                // 创建观察器以处理动态添加的链接
                var observer = new MutationObserver(function(mutations) {
                    modifyLinks();
                });
                
                // 开始观察整个档的变化
                observer.observe(document.documentElement, {
                    childList: true,
                    subtree: true
                });
            """)
            print("Successfully injected JavaScript code")
        except Exception as e:
            print(f"Error injecting JavaScript code: {e}")
