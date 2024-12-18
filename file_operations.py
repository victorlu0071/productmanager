import os
from tkinter import messagebox
import openpyxl
from PIL import Image
import io

def load_products(file_path):
    products = []
    product_dict = {}

    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2):
                product = {}
                for header, cell in zip(headers, row):
                    if header:
                        # 根据不同字段类型进行处理
                        if header == 'Cost':
                            # 处理成本字段
                            try:
                                value = float(cell.value) if cell.value is not None else 0.0
                                product[header] = f"¥{value:.2f}"
                            except (ValueError, TypeError):
                                product[header] = "¥0.00"
                        elif header == 'Code':
                            # 处理商品代码字段
                            try:
                                # 确保Code是整数
                                if isinstance(cell.value, int):
                                    value = cell.value
                                else:
                                    value = int(float(cell.value)) if cell.value is not None else 0
                                product[header] = value  # 直接存储整数，不转换为字符串
                            except (ValueError, TypeError):
                                product[header] = 0
                        elif header == 'StockQuantity':
                            # 处理库存数量字段
                            try:
                                value = int(float(cell.value)) if cell.value is not None else 0
                                product[header] = value
                            except (ValueError, TypeError):
                                product[header] = 0
                        elif header == 'Specs':
                            # 处理规格字段
                            product[header] = str(cell.value) if cell.value is not None else ""
                        else:
                            # 其他字段（包括Link）保持原始字符串
                            product[header] = str(cell.value) if cell.value is not None else ""
                
                products.append(product)
                if 'Code' in product:
                    try:
                        # 确保商品代码是整数
                        product_code = product['Code']
                        if isinstance(product_code, str):
                            product_code = int(float(product_code))
                        elif isinstance(product_code, float):
                            product_code = int(product_code)
                        product_dict[product_code] = product
                    except (ValueError, TypeError) as e:
                        print(f"处理商品代码出错: {e}")
                        continue
            
            wb.close()
        
        except FileNotFoundError:
            messagebox.showerror("Error", "The products file was not found.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred while loading products: {e}")

    return products, product_dict

def save_products(file_path, products):
    retries = 3
    for attempt in range(retries):
        try:
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 写入表头
            headers = ["Name", "Specs", "Cost", "Link", "Code", "LocationCode", "StockQuantity", "AddedDate"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, product in enumerate(products, 2):
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    value = product.get(header, "")
                    
                    # 根据字段类型设置单元格格式和值
                    if header == "Specs":
                        cell.value = str(value)
                    elif header == "Cost":
                        try:
                            if isinstance(value, str) and value.startswith('¥'):
                                value = float(value.replace('¥', '').strip())
                            else:
                                value = float(value) if value else 0.0
                        except (ValueError, TypeError):
                            value = 0.0
                        cell.value = value
                        cell.number_format = '¥#,##0.00'
                    
                    elif header == "Code":
                        try:
                            # 确保Code是整数
                            if isinstance(value, int):
                                cell.value = value
                            else:
                                cell.value = int(float(str(value))) if value else 0
                        except (ValueError, TypeError):
                            cell.value = 0
                        cell.number_format = '0'
                    
                    elif header == "StockQuantity":
                        try:
                            value = int(float(str(value))) if value else 0
                        except (ValueError, TypeError):
                            value = 0
                        cell.value = value
                        cell.number_format = '0'
                    
                    elif header == "LocationCode":
                        cell.value = str(value)
                        cell.number_format = '@'  # 文本格式
                    
                    else:
                        cell.value = str(value)
                
                # 设置行高以适应���片
                ws.row_dimensions[row].height = 150
                
                # 尝试添加图片
                try:
                    # 获取商品代码并确保是整数
                    product_code = product.get("Code", 0)
                    if isinstance(product_code, str):
                        product_code = int(float(product_code))
                    elif isinstance(product_code, float):
                        product_code = int(product_code)
                    
                    if product_code > 0:  # 只处理有效的商品代码
                        img_dir = os.path.join("product_images", str(product_code))
                        if os.path.exists(img_dir):
                            for i, img_file in enumerate(sorted(os.listdir(img_dir))):
                                if i >= 6:  # 最多显示6张图片
                                    break
                                if img_file.endswith(('.jpg', '.png')):
                                    img_path = os.path.join(img_dir, img_file)
                                    # 调整图片大小用于Excel显示
                                    img = Image.open(img_path)
                                    img.thumbnail((100, 100))
                                    temp_path = f"temp_img_{i}.png"
                                    img.save(temp_path)
                                    # 添加到Excel
                                    xl_img = openpyxl.drawing.image.Image(temp_path)
                                    col_letter = openpyxl.utils.get_column_letter(9 + i)  # 从第9列开始
                                    ws.add_image(xl_img, f"{col_letter}{row}")
                                    os.remove(temp_path)
                except Exception as e:
                    print(f"处理商品图片出错 (商品代码: {product_code}): {e}")
            
            # 保存文件
            wb.save(file_path)
            return True
            
        except PermissionError:
            if attempt < retries - 1:
                import time
                time.sleep(1)
            else:
                messagebox.showerror("Error", f"Permission denied: {file_path}. Ensure the file is not open in another program.")
                return False
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save products: {e}")
            return False
    
    return False
