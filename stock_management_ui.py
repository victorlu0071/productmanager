import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from PIL import Image
import os

class StockManagerUI:
    def __init__(self, products, product_dict, save_products):
        self.file_path = "products.xlsx"
        self.products = products
        # 确保product_dict的键都是整数类型
        self.product_dict = {int(k): v for k, v in product_dict.items()}
        self.save_products = save_products
        self.current_input_step = 0
        self.current_stock_entry = {}
        self.root = None
        self.mode = "out"  # 默认为出库模式
        self.timer = None  # 用于出库模式的计时器
        self.current_product = None  # 当前处理的产品
        self.current_stock_product = None  # 确保初始化时未选中产品
        self.setup_styles()  # 设置样式
        self.init_gui()

    def setup_styles(self):
        """设置UI样式"""
        # 定义颜色
        self.colors = {
            'primary': '#2196F3',  # 主色调
            'secondary': '#64B5F6',  # 次要色调
            'success': '#4CAF50',  # 成功色
            'warning': '#FFC107',  # 警告色
            'error': '#F44336',  # 错误色
            'background': '#F5F5F5',  # 背景色
            'text': '#212121',  # 文本色
            'light_text': '#757575'  # 浅色文本
        }

        # 创建自定义样式
        style = ttk.Style()
        style.theme_use('clam')

        # 配置Treeview样式
        style.configure("Treeview",
            background=self.colors['background'],
            foreground=self.colors['text'],
            fieldbackground=self.colors['background'],
            rowheight=30,
            font=('Microsoft YaHei UI', 10))
        
        style.configure("Treeview.Heading",
            background=self.colors['primary'],
            foreground="white",
            relief="flat",
            font=('Microsoft YaHei UI', 10, 'bold'))
        
        style.map("Treeview.Heading",
            background=[('active', self.colors['secondary'])])

        # 配置按钮样式
        style.configure("Primary.TButton",
            background=self.colors['primary'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10))
        
        style.map("Primary.TButton",
            background=[('active', self.colors['secondary'])])

        # 配置危险按钮样式
        style.configure("Danger.TButton",
            background=self.colors['error'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10))
        
        style.map("Danger.TButton",
            background=[('active', '#E57373')])

        # 配置选中按钮样式
        style.configure("Selected.TButton",
            background=self.colors['secondary'],
            foreground="white",
            padding=(20, 10),
            font=('Microsoft YaHei UI', 10, 'bold'))

        # 配置标签样式
        style.configure("Info.TLabel",
            background=self.colors['background'],
            foreground=self.colors['text'],
            font=('Microsoft YaHei UI', 10))

        # 配置输入框样式
        style.configure("Custom.TEntry",
            fieldbackground="white",
            padding=(5, 5))

    def save_to_excel(self):
        try:
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 写入表头
            headers = ["Name", "Cost", "Link", "Code", "LocationCode", "StockQuantity", "AddedDate"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, product in enumerate(self.products, 2):
                for col, header in enumerate(headers, 1):
                    # 确保Code作为数字存储
                    if header == "Code":
                        value = int(product.get(header, 0))
                    else:
                        value = product.get(header, "")
                    ws.cell(row=row, column=col, value=value)
                
                # 设置行以适应图片
                ws.row_dimensions[row].height = 150
                
                # 尝试添加图片
                try:
                    product_code = str(product.get("Code", ""))  # 转换为字符串用于文件路径
                    if product_code:
                        img_dir = os.path.join("product_images", product_code)
                        if os.path.exists(img_dir):
                            for i, img_file in enumerate(sorted(os.listdir(img_dir))):
                                if i >= 6:  # 最多显示6张图片
                                    break
                                if img_file.endswith(('.jpg', '.png')):
                                    img_path = os.path.join(img_dir, img_file)
                                    # 调整图片大小用于Excel显示
                                    img = Image.open(img_path)
                                    img.thumbnail((100, 100))
                                    temp_path = f"temp_img_{i}.png"
                                    img.save(temp_path)
                                    # 添加到Excel
                                    xl_img = openpyxl.drawing.image.Image(temp_path)
                                    col_letter = openpyxl.utils.get_column_letter(8 + i)  # 从第8列开始
                                    ws.add_image(xl_img, f"{col_letter}{row}")
                                    os.remove(temp_path)
                except Exception as e:
                    print(f"Error adding images for product {product_code}: {e}")
            
            # 保存文件
            wb.save(self.file_path)
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")

    def init_gui(self):
        try:
            self.root = tk.Tk()
            self.root.title("库存管理")
            
            # 确保窗口关闭时返回主菜单
            self.root.protocol("WM_DELETE_WINDOW", self.return_to_main_menu)
            
            # 设置窗口样式
            self.root.configure(bg=self.colors['background'])
            
            window_width = 1080
            window_height = 700
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            position_top = int((screen_height - window_height) / 2)
            position_left = int((screen_width - window_width) / 2)
            self.root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

            # 创建主框架
            main_frame = ttk.Frame(self.root)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

            # 添加模式切换按钮
            mode_frame = ttk.Frame(main_frame)
            mode_frame.pack(pady=5)
            
            self.in_button = ttk.Button(mode_frame, text="入库模式", 
                                      command=lambda: self.switch_mode("in"),
                                      style="Primary.TButton")
            self.in_button.pack(side=tk.LEFT, padx=5)
            
            self.out_button = ttk.Button(mode_frame, text="出库模式", 
                                       command=lambda: self.switch_mode("out"),
                                       style="Primary.TButton")
            self.out_button.pack(side=tk.LEFT, padx=5)

            # 输入区域
            input_frame = ttk.Frame(main_frame)
            input_frame.pack(fill=tk.X, pady=10)

            self.entry_label = ttk.Label(input_frame, 
                                       text="请输入要出库的产品编号:",
                                       style="Info.TLabel")
            self.entry_label.pack()

            self.entry_field = ttk.Entry(input_frame, 
                                       width=50,
                                       style="Custom.TEntry",
                                       font=('Microsoft YaHei UI', 10))
            self.entry_field.pack(pady=5)
            self.entry_field.bind('<Return>', self.handle_entry)
            self.root.bind('<Key>', self.focus_entry_on_keypress)

            # 添加倒计时标签
            self.timer_label = ttk.Label(input_frame, 
                                       text="",
                                       style="Info.TLabel")
            self.timer_label.pack()

            # Treeview区域
            tree_frame = ttk.Frame(main_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)

            # 创建Treeview的滚动条
            tree_scroll = ttk.Scrollbar(tree_frame)
            tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

            # Treeview
            self.tree = ttk.Treeview(tree_frame, 
                                   columns=("Code", "ProductName", "LocationCode", "StockQuantity"),
                                   show='headings',
                                   style="Treeview",
                                   yscrollcommand=tree_scroll.set)
            
            tree_scroll.config(command=self.tree.yview)

            # 设置列宽和表头
            column_headers = {
                "Code": "商品代码",
                "ProductName": "商品名称",
                "LocationCode": "库位码",
                "StockQuantity": "库存数量"
            }

            for col in column_headers:
                self.tree.heading(col, text=column_headers[col])
                self.tree.column(col, width=200, anchor='center')

            self.tree.pack(fill=tk.BOTH, expand=True)
            self.tree.bind('<Double-1>', self.edit_selected_item)

            # 按钮区域
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(pady=10)

            back_button = ttk.Button(button_frame, 
                                   text="返回主菜单", 
                                   command=self.return_to_main_menu,
                                   style="Primary.TButton")
            back_button.pack(pady=5)

            # 状态标签
            self.status_label = ttk.Label(main_frame, 
                                        text="", 
                                        anchor="w",
                                        style="Info.TLabel")
            self.status_label.pack(fill=tk.X, padx=10, pady=5)

            self.update_tree_view()
            self.update_mode_buttons()  # 这将设置出库按钮为按下状态

            self.root.mainloop()
        except tk.TclError as e:
            print(f"Error initializing GUI: {e}")
            self.return_to_main_menu()

    def update_tree_view(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for product_code, product in self.product_dict.items():
            self.tree.insert("", tk.END, values=(product_code, product.get("Name", ""), product.get("LocationCode", ""), product.get("StockQuantity", 0)))

    def edit_selected_item(self, event):
        selected_item = self.tree.selection()[0]
        column_index = self.tree.identify_column(event.x)
        column_number = int(column_index[1:]) - 1  # Convert to zero-indexed
        if column_number >= 0:
            entry_window = tk.Toplevel(self.root)
            entry_window.title("编辑数值")

            current_value = self.tree.item(selected_item, 'values')[column_number]
            entry_field = tk.Entry(entry_window, width=50)
            entry_field.insert(0, current_value)
            entry_field.pack(pady=10)

            def save_changes():
                new_value = entry_field.get()
                try:
                    values = list(self.tree.item(selected_item, 'values'))
                    product_code = int(values[0])  # 获取产品代码（第一列）
                    
                    if column_number == 3:  # StockQuantity
                        new_value = float(new_value)
                    
                    values[column_number] = new_value
                    self.tree.item(selected_item, values=values)
                    
                    # 更新product_dict和products
                    if product_code in self.product_dict:
                        if column_number == 1:  # Name
                            self.product_dict[product_code]["Name"] = new_value
                        elif column_number == 2:  # LocationCode
                            self.product_dict[product_code]["LocationCode"] = new_value
                        elif column_number == 3:  # StockQuantity
                            self.product_dict[product_code]["StockQuantity"] = new_value
                        
                        # 更新products列表
                        for product in self.products:
                            if int(product["Code"]) == product_code:
                                if column_number == 1:
                                    product["Name"] = new_value
                                elif column_number == 2:
                                    product["LocationCode"] = new_value
                                elif column_number == 3:
                                    product["StockQuantity"] = new_value
                                break
                    
                    self.save_products()  # 保存更改到文件
                    self.save_to_excel()  # 更新Excel文件
                    self.status_label.config(text="修改成功！")
                    entry_window.destroy()
                except ValueError:
                    self.status_label.config(text="错误: 请输入有效数值")
                    return

            save_button = tk.Button(entry_window, text="保存", command=save_changes)
            save_button.pack(pady=5)

    def focus_entry_on_keypress(self, event):
        if self.root.focus_get() != self.entry_field:
            self.entry_field.focus()
            self.entry_field.delete(0, tk.END)  # Clear existing text
            self.entry_field.insert(tk.END, event.char)

    def handle_entry(self, event):
        """处理输入框的回车事件"""
        if self.mode == "out":
            self.handle_checkout_entry()
        else:
            self.handle_stock_entry()

    def handle_checkout_entry(self):
        """处理出库模式的输入"""
        input_value = self.entry_field.get().strip()
        
        try:
            if not self.current_product:
                # 检查输入的是否是有效的产品代码
                product_code = int(input_value)  # 转换为整数
                print(f"Debug - Input code: {product_code}, Available codes: {list(self.product_dict.keys())}")  # 调试信息
                if product_code in self.product_dict:
                    # 检查库存是否为0
                    current_stock = float(self.product_dict[product_code].get("StockQuantity", 0))
                    if current_stock <= 0:
                        self.play_warning_sound()
                        messagebox.showwarning("警告", "该产品库存为0，无法出库！")
                        return
                    self.current_product = product_code
                    self.status_label.config(text=f"已选择商品: {self.product_dict[product_code].get('Name', '')}，请输入出库数量或等待自动出库1个")
                    self.start_checkout_timer()
                else:
                    self.status_label.config(text=f"错误: 未找到该产品编号 {product_code}")
            else:
                # 如果已经选择了产品，立即处理出库
                self.process_checkout()
                # 如果输入了数量，立即结束倒计时
                if input_value.isdigit():
                    self.cancel_timer()
        except ValueError:
            self.status_label.config(text="错误: 产品编号必须为数字")
        
        self.entry_field.delete(0, tk.END)

    def handle_stock_entry(self):
        """处理入库模式的输入"""
        user_input = self.entry_field.get().strip()
        self.entry_field.delete(0, tk.END)
        
        print("\n=== 入库操作开始 ===")
        print(f"用户输入: {user_input}")
        print(f"当前选中的产品: {self.current_stock_product}")
        if hasattr(self, 'current_stock_product') and self.current_stock_product:
            print(f"当前产品信息: {self.product_dict.get(self.current_stock_product)}")

        try:
            input_number = int(user_input)
            print(f"输入已转换为数字: {input_number}")
            
            # 检查是否是产品编号
            if input_number in self.product_dict:
                print(f"输入匹配到产品编号，产品信息: {self.product_dict[input_number]}")
                # 如果是已有的产品编号
                if hasattr(self, 'current_stock_product') and self.current_stock_product == input_number:
                    print("检测到重复扫描相同产品")
                    # 如果是相同产品，直接+1
                    self.update_stock_quantity(input_number, 1)
                    self.status_label.config(text=f"已增加库存 1 个")
                else:
                    print("检测到新产品扫描")
                    # 新的产品编号
                    self.current_stock_product = input_number
                    current_location = self.product_dict[input_number].get("LocationCode")
                    if current_location:
                        print(f"产品已有库位: {current_location}，等待验证库位")
                        self.entry_label.config(text="请扫描库位码以验证:")
                        self.status_label.config(text=f"已选择商品: {self.product_dict[input_number].get('Name', '')}，当前库位: {current_location}")
                    else:
                        print("产品无库位，等待输入库位")
                        self.entry_label.config(text="请扫描库位码:")
                        self.status_label.config(text=f"已选择商品: {self.product_dict[input_number].get('Name', '')}，请先扫描库位码")
                return
            
            # 如果已经选择了产品，检查是否是数量输入（1-99）
            if hasattr(self, 'current_stock_product') and self.current_stock_product:
                # 检查品是否有库位码
                if not self.product_dict[self.current_stock_product].get("LocationCode"):
                    print("尝试在设置库位码前输入数量")
                    self.play_warning_sound()
                    messagebox.showwarning("警告", "请先扫描库位码！")
                    return

                print(f"检查数量输入: {input_number}")
                if 1 <= input_number <= 99:
                    print(f"有效的数量输入: {input_number}")
                    # 输入1-99，需要减去之前已经增加的1个数量
                    self.update_stock_quantity(self.current_stock_product, input_number - 1)
                    self.current_stock_product = None
                    self.entry_label.config(text="请扫描产品编号:")
                    return
                else:
                    print(f"无效的数量输入: {input_number}")
                    # 无效的数量输入
                    self.status_label.config(text="错误: 数量必须在1-99之间")
                    return
            
            # 不是产品编号也不是数量，检查是否在等待库位编号
            if hasattr(self, 'current_stock_product') and self.current_stock_product:
                print(f"尝试处理库位编号: {input_number}")
                current_location = self.product_dict[self.current_stock_product].get("LocationCode")
                
                if current_location:
                    # 验证库位是否匹配
                    if str(input_number) != current_location:
                        print(f"库位不匹配，期望: {current_location}, 实际: {input_number}")
                        self.play_warning_sound()
                        messagebox.showwarning("警告", f"库位不匹配！\n正确库位: {current_location}\n扫描库位: {input_number}")
                        return
                    print("库位验证正确")
                else:
                    # 新增库位
                    # 检查库位码是否被其他产品使用
                    if self.check_location_code_exists(str(input_number)):
                        print("库位码已被其他产品使用")
                        self.play_warning_sound()
                        messagebox.showwarning("警告", "该库位码已被其他产品使用！")
                        return
                    print("保存新的库位码")
                    self.product_dict[self.current_stock_product]["LocationCode"] = str(input_number)
                
                # 增加库存
                self.update_stock_quantity(self.current_stock_product, 1)
                self.entry_label.config(text="请扫描产品编号:")
                self.save_products()  # 保存更新的库位信息
            else:
                print("无效输入，未选中产品")
                self.status_label.config(text="错误: 无效的输入")
                self.entry_label.config(text="请扫描产品编号:")
                
        except ValueError:
            print("输入不是数字，尝试处理为库位编号")
            # 输入的不是数字，可能是库位编号
            if hasattr(self, 'current_stock_product') and self.current_stock_product:
                current_location = self.product_dict[self.current_stock_product].get("LocationCode")
                
                if current_location:
                    # 验证库位是否匹配
                    if user_input != current_location:
                        print(f"库位不匹配，期望: {current_location}, 实际: {user_input}")
                        self.play_warning_sound()
                        messagebox.showwarning("警告", f"库位不匹配！\n正确库位: {current_location}\n扫描库位: {user_input}")
                        return
                    print("库位验证正确")
                else:
                    # 新增库位
                    # 检查库位码是否被其他产品使用
                    if self.check_location_code_exists(user_input):
                        print("库位码已被其他产品使用")
                        self.play_warning_sound()
                        messagebox.showwarning("警告", "该库位码已被其他产品使用！")
                        return
                    print("保存新的库位码（非数字）")
                    self.product_dict[self.current_stock_product]["LocationCode"] = user_input
                
                # 增加库存
                self.update_stock_quantity(self.current_stock_product, 1)
                self.entry_label.config(text="请扫描产品编号:")
                self.save_products()  # 保存更新的库位信息
            else:
                print("无效输入，未选中产品")
                self.status_label.config(text="错误: 请先扫描产品编号")
                self.entry_label.config(text="请扫描产品编号:")
        
        print("=== 入库操作结束 ===\n")

    def return_to_main_menu(self):
        try:
            if self.root:
                self.root.quit()
                self.root.destroy()
            import main
            main.ProductManager()
        except Exception as e:
            print(f"Error returning to main menu: {e}")
            # 确保在发生错误时也能退出
            if self.root:
                self.root.destroy()

    def switch_mode(self, mode):
        """切换入库/出库模式"""
        self.mode = mode
        self.current_input_step = 0
        self.current_stock_entry = {}
        self.current_product = None
        self.current_stock_product = None  # 重置选中的产品
        self.entry_field.delete(0, tk.END)
        self.cancel_timer()
        self.timer_label.config(text="")
        
        if mode == "in":
            self.entry_label.config(text="请扫描产品编号:")
        else:
            self.entry_label.config(text="请输入要出库的产品编号:")
        
        self.update_mode_buttons()

    def update_mode_buttons(self):
        """更新模式按钮的状态"""
        if self.mode == "in":
            self.in_button.configure(style="Selected.TButton")
            self.out_button.configure(style="Primary.TButton")
        else:
            self.in_button.configure(style="Primary.TButton")
            self.out_button.configure(style="Selected.TButton")

    def cancel_timer(self):
        """取消当前的计时器"""
        if self.timer:
            self.root.after_cancel(self.timer)
            self.timer = None

    def start_checkout_timer(self, seconds=10):
        """开始出库倒计时"""
        self.cancel_timer()
        self.countdown(seconds)

    def countdown(self, seconds):
        """倒计时处理"""
        if seconds > 0:
            self.timer_label.config(text=f"等待输入数量: {seconds}秒")
            self.timer = self.root.after(1000, lambda: self.countdown(seconds - 1))
        else:
            self.timer_label.config(text="")
            self.process_checkout()

    def process_checkout(self):
        """处理出库操作"""
        if not self.current_product:
            return

        # 获取输入框中的值
        input_value = self.entry_field.get().strip()
        
        try:
            # 检查是否输入了新的有效产品代码
            try:
                product_code = int(input_value)
                if product_code in self.product_dict:
                    # 检查当前产品是否有库存
                    current_stock = float(self.product_dict[self.current_product].get("StockQuantity", 0))
                    if current_stock <= 0:
                        self.play_warning_sound()
                        messagebox.showwarning("警告", "当前产品库存为0，无法出库！")
                        return
                    # 如果输入了新的产品代码，当前产品减1
                    self.update_stock_quantity(self.current_product, -1)
                    # 开始处理新的产品
                    self.current_product = product_code
                    self.start_checkout_timer()
                    return
            except ValueError:
                pass  # 不是有效的产品代码，继续处理数量
            
            # 检查是否输入了数量
            if input_value.isdigit() and len(input_value) <= 3:
                quantity = int(input_value)
                if quantity > 0:
                    # 检查库存是否足够
                    current_stock = float(self.product_dict[self.current_product].get("StockQuantity", 0))
                    if current_stock < quantity:
                        self.play_warning_sound()
                        messagebox.showwarning("警告", f"库存不足！\n当前库存: {current_stock}\n要出库数量: {quantity}")
                        return
                    self.update_stock_quantity(self.current_product, -quantity)
                    # 输入数量后立即结束倒计时
                    self.cancel_timer()
            else:
                # 检查当前产品是否有库存
                current_stock = float(self.product_dict[self.current_product].get("StockQuantity", 0))
                if current_stock <= 0:
                    self.play_warning_sound()
                    messagebox.showwarning("警告", "当前产品库存为0，无法出库！")
                    return
                # 如果没有输入数量或输入无效，默认减1
                self.update_stock_quantity(self.current_product, -1)
        except Exception as e:
            self.status_label.config(text=f"出库处理错误: {str(e)}")
        
        # 重置状态
        self.current_product = None
        self.entry_field.delete(0, tk.END)
        self.timer_label.config(text="")
        self.entry_label.config(text="请输入要出库的产品编号:")

    def update_stock_quantity(self, product_code, change):
        """更新库存数量"""
        if product_code in self.product_dict:
            product = self.product_dict[product_code]
            current_quantity = float(product.get("StockQuantity", 0))
            new_quantity = max(0, current_quantity + change)  # 确保数量小于0
            product["StockQuantity"] = new_quantity
            
            # 更新Excel文件
            self.save_products()
            self.update_tree_view()
            
            # 更新状态消息
            operation = "出库" if change < 0 else "入库"
            self.status_label.config(text=f"成功{operation} {abs(change)} 个商品 (Code: {product_code})")

    def play_warning_sound(self):
        """播放警告音"""
        self.root.bell()  # 使用系统提示音

    def check_location_code_exists(self, location_code):
        """检查库位码是否已存在"""
        for code, product in self.product_dict.items():
            if product.get("LocationCode") == location_code and code != self.current_stock_product:
                return True
        return False
